import os
import openpyxl
from openpyxl.styles import PatternFill


def read_txt_files(raw_label_path):
    txt_files = [file for file in os.listdir(raw_label_path) if file.endswith(".txt")]
    data = []
    for txt_file in txt_files:
        with open(os.path.join(raw_label_path, txt_file), "r") as f:
            lines = f.readlines()
            data.append([line.strip().split(":") for line in lines])
    assert len(set(len(d) for d in data)) == 1, "所有txt文件的行数不一致"
    return data, txt_files


def is_float(value):
    try:
        float(value)
        return True
    except ValueError:
        return False


def generate_excel(data, txt_files, output_xlsx):
    number_of_worker = len(data)
    
    wb = openpyxl.Workbook()
    ws = wb.active

    header = ['Filenames']
    header.extend(f'{txt_file} Performance' for txt_file in txt_files)
    header.append('AVG_Performance')
    header.extend(f'{txt_file} Placement' for txt_file in txt_files)
    header.append('AVG_Placement')
    header.append('note')
    ws.append(header)

    for i in range(len(data[0])):
        row = [data[0][i][0]]  # 路径名
        labels1 = []
        labels2 = []
        notes = []

        for j in range(len(txt_files)):
            label = data[j][i][1]
            if "delete" in label:
                labels1.append("delete")
                labels2.append("delete")
            else:
                label_parts = label.split(",")
                labels1.append(label_parts[0])
                labels2.append(label_parts[1] if len(label_parts) > 1 else "delete")
            
            if len(label_parts) > 2:
                notes.append(label_parts[2])
            else:
                notes.append("")

        row.extend(labels1)
        valid_labels1 = [label for label in labels1 if label == "delete" or is_float(label)]
        row.append(round(sum(0 if label == "delete" else float(label) for label in valid_labels1) / len(valid_labels1)) if valid_labels1 else "")  # Performance的平均数
        
        row.extend(labels2)
        valid_labels2 = [label for label in labels2 if label == "delete" or is_float(label)]
        row.append(round(sum(0 if label == "delete" else float(label) for label in valid_labels2) / len(valid_labels2)) if valid_labels2 else "")  # Placement的平均数
        
        row.append("\n".join(note for note in notes if note))  # 所有的note

        ws.append(row)

    wb.save(output_xlsx)


def highlight_rows(output_xlsx, txt_files):
    wb = openpyxl.load_workbook(output_xlsx)
    ws = wb.active

    total = 0
    count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=len(txt_files)*2+3):
        row_values = [cell.value for cell in row]
        #print(row_values)
        labels1 = row_values[0:len(txt_files)]
        #labels2 = row_values[len(txt_files)+2:-1]
        labels2 = row_values[4:7]
        
        #print("label2:", labels2)
        total += 1
        
        if ("1" in labels2 and "4" in labels2):
            count += 1
        # Focus on the Placement
        #if ("1" in labels2 and "4" in labels2):
            for cell in ws[row[0].row][:len(txt_files)*2+3]:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
    print(total)
    print(count)

    wb.save(output_xlsx)
    
def generate_final_gt_file(output_xlsx, gt_file):
    if os.path.exists(gt_file):
        os.remove(gt_file)

    wb = openpyxl.load_workbook(output_xlsx)
    ws = wb.active

    with open(gt_file, 'w') as f:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            filename = row[0].value
            avg_placement = row[-2].value

            if avg_placement == "Review":
                raise Exception("发现有Review标签,请检查数据!")
            elif avg_placement == "delete":
                continue
            else:
                new_label = avg_placement
                f.write(f"{filename}:{new_label}\n")

    with open(gt_file, 'r') as f:
        num_lines = sum(1 for _ in f)

    print(f"gt文件总行数为: {num_lines}")

def generate_excel_binary(data, txt_files, output_xlsx):
    number_of_worker = len(data)
        
    wb = openpyxl.Workbook()
    ws = wb.active

    header = ['Filenames']
    header.extend(f'{txt_file} Performance' for txt_file in txt_files)
    header.append('AVG_Performance')
    header.extend(f'{txt_file} Placement' for txt_file in txt_files)
    header.append('AVG_Placement')
    header.append('note')
    ws.append(header)

    for i in range(len(data[0])):
        row = [data[0][i][0]]  # 路径名
        labels1 = []
        labels2 = []
        notes = []

        for j in range(len(txt_files)):
            try:
                label = data[j][i][1].strip()
                if "delete" in label:
                    labels1.append("delete")
                    labels2.append("delete")
                else:
                    label_parts = label.split(",")
                    labels1.append(True if int(label_parts[0]) in [3, 4] else False)
                    labels2.append(True if int(label_parts[1]) in [3, 4] else False)

                if len(label_parts) > 2:
                    notes.append(label_parts[2])
                else:
                    notes.append("")
            except Exception as e:
                print(e)
                print(i)
                

        row.extend(labels1)

        true_count = labels1.count(True)
        false_count = labels1.count(False)
        delete_count = labels1.count("delete")
        if delete_count > number_of_worker / 2:
            row.append("delete")
        elif true_count > false_count:
            row.append(True)
        elif false_count > true_count:
            row.append(False)
        else:
            row.append("Review")
            
        row.extend(labels2)

        true_count = labels2.count(True)
        false_count = labels2.count(False)
        delete_count = labels2.count("delete")
        if delete_count > number_of_worker / 2:
            row.append("delete")
        elif (true_count > false_count) and (true_count + false_count == number_of_worker):
            row.append(True)
        elif (false_count > true_count) and (true_count + false_count == number_of_worker):
            row.append(False)
        else:
            row.append("Review")

        row.append("\n".join(note for note in notes if note))  # 所有的note
        ws.append(row)

    wb.save(output_xlsx)

def highlight_rows_binary(output_xlsx, txt_files):
    wb = openpyxl.load_workbook(output_xlsx)
    ws = wb.active

    num_review = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        row_values = [cell.value for cell in row]

        if row_values[-2] == "Review":
            num_review += 1
            for cell in ws[row[0].row][1:ws.max_column]:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
    print(f"需要review的文件数: {num_review}")
    wb.save(output_xlsx)
    return num_review

def generate_final_gt_file_binary(output_xlsx, gt_file):
    if os.path.exists(gt_file):
        os.remove(gt_file)

    wb = openpyxl.load_workbook(output_xlsx)
    ws = wb.active

    with open(gt_file, 'w') as f:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            filename = row[0].value
            avg_placement = row[-2].value

            if avg_placement == "Review":
                raise Exception("发现有Review标签,请检查数据!")
            elif avg_placement == "delete":
                continue
            else:
                new_label = 1 if avg_placement else 0
                f.write(f"{filename}:{new_label}\n")

    with open(gt_file, 'r') as f:
        num_lines = sum(1 for _ in f)

    print(f"gt文件总行数为: {num_lines}")
        

def main():
    raw_label_path = "raw_label"
    #output_xlsx = "output.xlsx"
    output_xlsx = "output_binary.xlsx"
    gt_file = "gt_final.txt"
    
    if os.path.exists(output_xlsx):
        os.remove(output_xlsx)

    data, txt_files = read_txt_files(raw_label_path)
    
    # For 4-Classes
    #generate_excel(data, txt_files, output_xlsx)
    #highlight_rows(output_xlsx, txt_files)
    #generate_final_gt_file(output_xlsx, gt_file)
    
    # For Binary-Classes
    generate_excel_binary(data, txt_files, output_xlsx)
    num_review = highlight_rows_binary(output_xlsx, txt_files)
    
    print("Excel file generated")
    
    if num_review == 0:
        generate_final_gt_file_binary(output_xlsx, gt_file)
        print("New gt file generated")


if __name__ == "__main__":
    main()